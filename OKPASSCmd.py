import os
import shutil
from datetime import datetime
import openpyxl

# 配置常量
MAIN_FILE = 'records.xlsx'
TEMPLATE_PATH = os.path.join('template', 'template.xlsx')
OLD_RECORDS_DIR = 'OldRecords'
MAX_RECORDS = 99

def init_environment():
    """初始化运行环境"""
    try:
        os.makedirs(OLD_RECORDS_DIR, exist_ok=True)
        os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)
        
        # 如果主文件不存在则从模板创建
        if not os.path.exists(MAIN_FILE):
            shutil.copy(TEMPLATE_PATH, MAIN_FILE)
            print(f"新建主文件：{MAIN_FILE}")
            
        return True
    except Exception as e:
        print(f"初始化失败：{str(e)}")
        return False

def get_current_workbook():
    """获取当前工作簿"""
    try:
        if not os.path.exists(MAIN_FILE):
            shutil.copy(TEMPLATE_PATH, MAIN_FILE)
        wb = openpyxl.load_workbook(MAIN_FILE)
        return wb, wb['Record']
    except Exception as e:
        print(f"加载工作簿失败：{str(e)}")
        return None, None

def show_record_count():
    """显示当前记录数"""
    try:
        wb, ws = get_current_workbook()
        count = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
        print(f"\n当前记录数：{count}/{MAX_RECORDS}")
        return count
    except:
        return 0

def add_record():
    """添加新记录"""
    try:
        wb, ws = get_current_workbook()
        
        record = [
            input("输入日期 (DD/MM/YYYY): "),
            input("输入时间 (XX:XX - XX:XX): "),
            input("输入名称: ")
        ]
        
        # 查找第一个空行（从第二行开始）
        row = 2
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        
        # 写入数据
        for col, value in enumerate(record[:3], 1):
            ws.cell(row=row, column=col, value=value)
        
        wb.save(MAIN_FILE)
        print("记录添加成功！")
        
        # 检查记录数
        current_count = sum(1 for r in ws.iter_rows(min_row=2) if r[0].value)
        if current_count >= MAX_RECORDS - 5:
            print(f"提示：剩余空间 {MAX_RECORDS - current_count} 条")
        return True
    except Exception as e:
        print(f"添加记录失败：{str(e)}")
        return False

def modify_record():
    """修改记录"""
    try:
        wb, ws = get_current_workbook()
        
        # 搜索记录
        search_name = input("请输入要修改的记录名称：").strip()
        matches = []
        
        # 遍历所有记录行（从第二行开始）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            name_cell = row[2]  # 名称在第三列
            if name_cell.value and name_cell.value.lower() == search_name.lower():
                matches.append({
                    "row": row_idx,
                    "date": row[0].value,
                    "time": row[1].value,
                    "name": name_cell.value
                })
        
        if not matches:
            print("未找到匹配记录")
            return False
            
        # 显示搜索结果
        print(f"\n找到 {len(matches)} 条匹配记录：")
        for idx, match in enumerate(matches, 1):
            print(f"{idx}. {match['date']} | {match['time']} | {match['name']}")
        
        # 选择要修改的记录
        while True:
            try:
                choice = int(input("\n请选择要修改的记录编号（0取消）："))
                if 0 <= choice <= len(matches):
                    break
            except ValueError:
                pass
            print("无效输入，请重新输入")
            
        if choice == 0:
            print("操作已取消")
            return False
            
        selected = matches[choice-1]
        
        # 获取新数据
        print(f"\n正在修改记录：{selected['name']}")
        new_date = input(f"新日期 [{selected['date']}]（直接回车保持不变）：") 
        new_time = input(f"新时间 [{selected['time']}]（直接回车保持不变）：")
        
        # 更新数据
        if new_date:
            ws.cell(row=selected['row'], column=1, value=new_date)
        if new_time:
            ws.cell(row=selected['row'], column=2, value=new_time)
        
        wb.save(MAIN_FILE)
        print("记录修改成功！")
        return True
    except Exception as e:
        print(f"修改失败：{str(e)}")
        return False

def archive_and_reset():
    """归档当前文件并创建新文件"""
    try:
        if not os.path.exists(MAIN_FILE):
            print("当前文件不存在")
            return False
            
        # 生成归档文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"archive_{timestamp}.xlsx"
        dest_path = os.path.join(OLD_RECORDS_DIR, archive_name)
        
        # 移动当前文件到归档目录
        shutil.move(MAIN_FILE, dest_path)
        
        # 创建新文件
        shutil.copy(TEMPLATE_PATH, MAIN_FILE)
        print(f"已归档旧文件并创建新文件：{MAIN_FILE}")
        return True
    except Exception as e:
        print(f"归档失败：{str(e)}")
        return False

def main_menu():
    """主菜单系统"""
    if not init_environment():
        return
    
    while True:
        # 显示记录数
        current_count = show_record_count()
        
        # 强制归档检查
        if current_count >= MAX_RECORDS:
            print("\n记录已满，必须归档！")
            archive_and_reset()
            continue
                
        # 主菜单
        print("\n==== 记录管理系统 ====")
        print("1. 添加新记录")
        print("2. 修改现有记录")
        print("3. 手动归档文件")
        print("4. 退出系统")
        
        choice = input("请选择操作 (1-4): ")
        
        if choice == '1':
            add_record()
        elif choice == '2':
            modify_record()
        elif choice == '3':
            if input("确认归档当前文件？(y/n) ").lower() == 'y':
                archive_and_reset()
        elif choice == '4':
            print("程序已关闭，按任意键退出")
            break
        else:
            print("无效输入，请重新选择")

if __name__ == "__main__":
    main_menu()
