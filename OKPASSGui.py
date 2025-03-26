import os 
import shutil
from datetime import datetime
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from tkcalendar import DateEntry  # 新增日历组件

class RecordManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("记录管理系统")
        self.root.geometry("800x600")
        
        # 常量配置
        self.MAIN_FILE = 'records.xlsx'
        self.TEMPLATE_PATH = os.path.join('template', 'template.xlsx')
        self.OLD_RECORDS_DIR = 'OldRecords'
        self.MAX_RECORDS = 99
        
        # 初始化环境
        self.init_environment()
        
        # 创建GUI组件
        self.create_widgets()
        
        # 初始加载数据
        self.update_record_count()

    def init_environment(self):
        """初始化运行环境"""
        try:
            os.makedirs(self.OLD_RECORDS_DIR, exist_ok=True)
            os.makedirs(os.path.dirname(self.TEMPLATE_PATH), exist_ok=True)
            
            if not os.path.exists(self.MAIN_FILE):
                shutil.copy(self.TEMPLATE_PATH, self.MAIN_FILE)
                self.log_message("新建主文件：records.xlsx")
        except Exception as e:
            messagebox.showerror("初始化错误", str(e))
            self.root.destroy()

    def create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(expand=True, fill=tk.BOTH)
        
        # 记录数显示
        self.count_label = ttk.Label(main_frame, text="当前记录数：0/99", font=('Arial', 12))
        self.count_label.pack(pady=10)
        
        # 功能按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="添加记录", command=self.show_add_dialog, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="修改记录", command=self.show_modify_dialog, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="归档文件", command=self.archive_file, width=15).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="退出系统", command=self.root.quit, width=15).pack(side=tk.LEFT, padx=10)
        
        # 日志显示
        self.log_text = tk.Text(main_frame, height=10, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=10)

    def update_record_count(self):
        """更新记录数显示"""
        try:
            wb, ws = self.get_current_workbook()
            count = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            self.count_label.config(text=f"当前记录数：{count}/{self.MAX_RECORDS}")
            
            if count >= self.MAX_RECORDS:
                messagebox.showwarning("记录已满", "必须立即归档！")
                self.archive_file()
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def get_current_workbook(self):
        """获取当前工作簿和工作表对象"""
        try:
            if not os.path.exists(self.MAIN_FILE):
                shutil.copy(self.TEMPLATE_PATH, self.MAIN_FILE)
            wb = openpyxl.load_workbook(self.MAIN_FILE)
            return wb, wb['Record']  # 返回工作表对象而不是名称
        except Exception as e:
            messagebox.showerror("错误", f"加载工作簿失败：{str(e)}")
            return None, None

    def show_add_dialog(self):
        """显示添加记录对话框（使用日历和时间选择控件）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加新记录")
        dialog.grab_set()
        
        # 日期选择
        ttk.Label(dialog, text="日期 (DD/MM/YYYY):").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        date_entry = DateEntry(dialog, date_pattern='dd/MM/yyyy', width=22)
        date_entry.grid(row=0, column=1, padx=10, pady=5)
        
        # 时间选择控件（开始时间和结束时间）
        ttk.Label(dialog, text="开始时间:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        start_frame = ttk.Frame(dialog)
        start_frame.grid(row=1, column=1, padx=10, pady=5)
        start_hour = ttk.Combobox(start_frame, width=3, values=[str(i).zfill(2) for i in range(24)], state='readonly')
        start_hour.current(8)
        start_hour.pack(side=tk.LEFT)
        ttk.Label(start_frame, text=":").pack(side=tk.LEFT)
        start_minute = ttk.Combobox(start_frame, width=3, values=[str(i).zfill(2) for i in range(60)], state='readonly')
        start_minute.current(0)
        start_minute.pack(side=tk.LEFT)
        
        ttk.Label(dialog, text="结束时间:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        end_frame = ttk.Frame(dialog)
        end_frame.grid(row=2, column=1, padx=10, pady=5)
        end_hour = ttk.Combobox(end_frame, width=3, values=[str(i).zfill(2) for i in range(24)], state='readonly')
        end_hour.current(9)
        end_hour.pack(side=tk.LEFT)
        ttk.Label(end_frame, text=":").pack(side=tk.LEFT)
        end_minute = ttk.Combobox(end_frame, width=3, values=[str(i).zfill(2) for i in range(60)], state='readonly')
        end_minute.current(0)
        end_minute.pack(side=tk.LEFT)
        
        # 名称输入
        ttk.Label(dialog, text="名称:").grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
        name_entry = ttk.Entry(dialog, width=25)
        name_entry.grid(row=3, column=1, padx=10, pady=5)
        
        def on_submit():
            try:
                # 获取日期（已是正确格式 dd/MM/yyyy）
                date_str = date_entry.get()
                # 获取时间，并构造字符串 格式为 "HH:MM - HH:MM"
                t1 = f"{start_hour.get()}:{start_minute.get()}"
                t2 = f"{end_hour.get()}:{end_minute.get()}"
                time_str = f"{t1} - {t2}"
                
                name = name_entry.get().strip()
                if not date_str or not time_str or not name:
                    raise ValueError("所有字段必须填写")
                
                self.add_record([date_str, time_str, name])
                dialog.destroy()
                self.update_record_count()
            except Exception as e:
                messagebox.showerror("错误", str(e))
        
        ttk.Button(dialog, text="提交", command=on_submit).grid(row=4, columnspan=2, pady=10)

    def add_record(self, record):
        """添加新记录"""
        try:
            wb, ws = self.get_current_workbook()
            
            # 查找第一个空行（从第二行开始）
            row = 2
            while ws.cell(row=row, column=1).value is not None:
                row += 1
            
            # 写入数据
            for col, value in enumerate(record, 1):
                ws.cell(row=row, column=col, value=value)
            
            wb.save(self.MAIN_FILE)
            self.log_message(f"已添加记录：{record[2]}")
            
            # 检查记录数
            current_count = sum(1 for r in ws.iter_rows(min_row=2) if r[0].value)
            if current_count >= self.MAX_RECORDS - 5:
                self.log_message(f"提示：剩余空间 {self.MAX_RECORDS - current_count} 条")
        except Exception as e:
            raise Exception(f"添加失败：{str(e)}")

    def show_modify_dialog(self):
        """显示修改记录对话框"""
        search_term = simpledialog.askstring("搜索记录", "请输入要修改的记录名称：")
        if not search_term:
            return
            
        matches = self.search_records(search_term)
        if not matches:
            messagebox.showinfo("无结果", "未找到匹配记录")
            return
            
        self.show_selection_dialog(matches)

    def search_records(self, search_term):
        """搜索记录"""
        try:
            wb, ws = self.get_current_workbook()
            matches = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                name_cell = row[2]  # 名称在第三列
                if name_cell.value and search_term.lower() in name_cell.value.lower():
                    matches.append({
                        'row': row_idx,
                        'date': row[0].value,
                        'time': row[1].value,
                        'name': name_cell.value
                    })
            return matches
        except Exception as e:
            messagebox.showerror("错误", str(e))
            return []

    def show_selection_dialog(self, matches):
        """显示选择记录对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("选择要修改的记录")
        dialog.grab_set()
        
        # 创建表格显示结果
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        columns = ("#1", "#2", "#3")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=6)
        
        # 设置表头
        tree.heading("#1", text="日期")
        tree.heading("#2", text="时间")
        tree.heading("#3", text="名称")
        
        # 设置列宽
        tree.column("#1", width=120, anchor=tk.CENTER)
        tree.column("#2", width=150, anchor=tk.CENTER)
        tree.column("#3", width=200, anchor=tk.W)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 填充数据
        for match in matches:
            tree.insert("", "end", values=(match['date'], match['time'], match['name']))
        
        # 布局组件
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 选择按钮
        def on_select():
            selected_item = tree.focus()
            if selected_item:
                item_data = tree.item(selected_item)
                selected_index = tree.index(selected_item)
                self.show_edit_dialog(matches[selected_index])
                dialog.destroy()
        
        ttk.Button(dialog, text="选择", command=on_select).pack(pady=10)

    def show_edit_dialog(self, record):
        """显示编辑对话框（使用日历和时间选择控件）"""
        dialog = tk.Toplevel(self.root)
        dialog.title("编辑记录")
        dialog.grab_set()
        
        # 日期选择，解析已有日期
        ttk.Label(dialog, text="日期:").grid(row=0, column=0, padx=10, pady=5, sticky=tk.E)
        try:
            init_date = datetime.strptime(record['date'], "%d/%m/%Y")
        except Exception:
            init_date = datetime.now()
        date_entry = DateEntry(dialog, date_pattern='dd/MM/yyyy', width=22)
        date_entry.set_date(init_date)
        date_entry.grid(row=0, column=1, padx=10, pady=5)
        
        # 时间选择（解析已有时间，格式 “HH:MM - HH:MM”）
        ttk.Label(dialog, text="开始时间:").grid(row=1, column=0, padx=10, pady=5, sticky=tk.E)
        start_frame = ttk.Frame(dialog)
        start_frame.grid(row=1, column=1, padx=10, pady=5)
        start_hour = ttk.Combobox(start_frame, width=3, values=[str(i).zfill(2) for i in range(24)], state='readonly')
        start_minute = ttk.Combobox(start_frame, width=3, values=[str(i).zfill(2) for i in range(60)], state='readonly')
        # 结束时间
        ttk.Label(dialog, text="结束时间:").grid(row=2, column=0, padx=10, pady=5, sticky=tk.E)
        end_frame = ttk.Frame(dialog)
        end_frame.grid(row=2, column=1, padx=10, pady=5)
        end_hour = ttk.Combobox(end_frame, width=3, values=[str(i).zfill(2) for i in range(24)], state='readonly')
        end_minute = ttk.Combobox(end_frame, width=3, values=[str(i).zfill(2) for i in range(60)], state='readonly')
        
        # 解析原时间数据
        try:
            t1, t2 = record['time'].split(" - ")
            sh, sm = t1.split(":")
            eh, em = t2.split(":")
        except Exception:
            sh, sm, eh, em = "08", "00", "09", "00"
        
        start_hour.set(sh)
        start_minute.set(sm)
        end_hour.set(eh)
        end_minute.set(em)
        start_hour.pack(side=tk.LEFT)
        ttk.Label(start_frame, text=":").pack(side=tk.LEFT)
        start_minute.pack(side=tk.LEFT)
        end_hour.pack(in_=end_frame, side=tk.LEFT)
        ttk.Label(end_frame, text=":").pack(side=tk.LEFT)
        end_minute.pack(in_=end_frame, side=tk.LEFT)
        
        def on_save():
            try:
                new_date = date_entry.get()
                t1 = f"{start_hour.get()}:{start_minute.get()}"
                t2 = f"{end_hour.get()}:{end_minute.get()}"
                new_time = f"{t1} - {t2}"
                
                if not new_date or not new_time:
                    raise ValueError("日期和时间不能为空")
                
                self.update_record(record['row'], new_date, new_time)
                dialog.destroy()
                self.update_record_count()
            except Exception as e:
                messagebox.showerror("错误", str(e))
        
        ttk.Button(dialog, text="保存", command=on_save).grid(row=3, columnspan=2, pady=10)

    def update_record(self, row, new_date, new_time):
        """更新记录"""
        try:
            wb, ws = self.get_current_workbook()
            
            ws.cell(row=row, column=1, value=new_date)
            ws.cell(row=row, column=2, value=new_time)
            
            wb.save(self.MAIN_FILE)
            self.log_message(f"已更新记录：第{row}行")
        except Exception as e:
            raise Exception(f"更新失败：{str(e)}")

    def archive_file(self):
        """归档文件"""
        if messagebox.askyesno("确认归档", "确定要归档当前文件并创建新文件吗？"):
            try:
                # 生成归档文件名
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                archive_name = f"archive_{timestamp}.xlsx"
                dest_path = os.path.join(self.OLD_RECORDS_DIR, archive_name)
                
                # 移动文件
                shutil.move(self.MAIN_FILE, dest_path)
                
                # 创建新文件
                shutil.copy(self.TEMPLATE_PATH, self.MAIN_FILE)
                
                self.log_message(f"已归档文件：{archive_name}")
                self.update_record_count()
                messagebox.showinfo("成功", "文件归档完成，已创建新文件")
            except Exception as e:
                messagebox.showerror("错误", str(e))

    def log_message(self, message):
        """记录日志信息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    app = RecordManagerApp(root)
    root.mainloop()
